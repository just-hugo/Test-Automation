import xlwt
from xlrd import open_workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np

#workbook = open_workbook(r'C:\Users\CheyanneHewitt\PycharmProjects\UnitedFlightBooking\airportCodes.xlsx')

workbook = load_workbook(r'C:\Users\CheyanneHewitt\PycharmProjects\UnitedFlightBooking\airportCodes.xlsx')
#print(workbook.sheetnames)

sheet = workbook.active
#sheet = workbook.sheet_by_name('airportCodes')
airports = []

for row in sheet.rows:
    for column in sheet.columns:
        d = {row : column}
        airports.append(d)

for i in range(len(airports)):
    print(airports(i))


